# from openpyxl import Workbook
from openpyxl import load_workbook
# from Tkinter import *
# from tkFileDialog import askopenfilename
from tkinter import filedialog
from tkinter import *
import tkinter as tk

class ListItem:
    def __init__(self, code, name):
        self.code = code
        self.name = name


class Posting:
    def __init__(self, row):
        self.deptCode = row[0].value
        self.deptName = row[1].value
        self.nominalCode = row[2].value
        self.nominalName = row[3].value
        self.date = row[4].value
        self.narrative = row[5].value
        self.blank = row[6].value
        self.transValue = row[7].value
        self.cat = row[8].value


wb = load_workbook('data.xlsx')
ws = wb['Sheet1']


def printSheetTitles(wb):
    for sheet in wb:
        print(sheet.title)


def getRow(ws, rowno):
    newRow = ws[rowno]
    # for item in newRow:
    #     print(item.value)
    # print(newRow)
    return newRow


def printvalues(ws):
    for row in ws.values:
        for value in row:
            print(value)


def categorise(ws):
    for i in range(1, ws.max_row):
        cell = ws.cell(i, 3)
        if (cell.value == 7003) or (cell.value == 7006) or (cell.value == 7007):
            ws.cell(i, 9, 'Staff Costs')
        if cell.value == 5032:
            ws.cell(i, 9, 'Equipment')
        if (cell.value == 7307) or (cell.value == 7311) or (cell.value == 7312):
            ws.cell(i, 9, 'Travel')


def searchByNominal(entryList, nominalCode):
    filteredList = []
    for item in entryList:
        if item.nominalCode == nominalCode:
            filteredList.append(item)
    return filteredList


def save(wb, name):
    wb.save(name)


def printrows(ws):
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
        print(row)


def sumcat(ws, cat):
    catsum = 0
    for i in range(1, ws.max_row):
        cell = ws.cell(i, 9)
        if cell.value == cat:
            cell = ws.cell(i, 8)
            catsum = catsum + cell.value
    string = (cat + ' Total : ' + str(catsum))
#     print(string)
    return string

def sumdept(ws, dept):
    deptsum = 0
    for i in range(1, ws.max_row):
        cell = ws.cell(i, 1)
        if cell.value == dept:
            cell = ws.cell(i, 8)
            name = ws.cell(i, 2).value
            deptsum += cell.value
    string = (name + ' Total : ' + str(deptsum))
    return string


def getList(ws, codeCol, nameCol):
    list = []
    initCode = ws.cell(1, codeCol).value
    initName = ws.cell(1, nameCol).value
    list.append(ListItem(initCode, initName))
    for i in range(2, ws.max_row):
        itemCode = ws.cell(i, codeCol).value
        itemName = ws.cell(i, nameCol).value
        # newDept = Department(deptCode, deptName)
        # deptList.append(newDept)
        for item in list:
            found = False
            if item.code == itemCode:
                found = True
        if found == False:
            newItem = ListItem(itemCode, itemName)
            list.append(newItem)
    return list


def printList(list):
    print("No. of List Items: " + str(list.__len__()))
    for item in list:
        print(str(item.code) + " " + item.name)


def writeList(ws, list, row, column, sheetname):
    newWs = wb.create_sheet(sheetname)
    i = 0
    for item in list:
        newWs.cell(row + i, column).value = item.code
        newWs.cell(row + i, column + 1).value = item.name
        i += 1


def createEntries(ws):
    entryList = []
    for i in range (1, ws.max_row):
        newEntry = Posting(getRow(ws, i))
        entryList.append(newEntry)
    print(entryList.__len__())
    return entryList


def writeEntryList(entryList, col1, sheetname):
    newWs = wb.create_sheet(sheetname)
    for i in range(1, entryList.__len__()):
        newWs.cell(i, 1).value = entryList[i].nominalCode
        newWs.cell(i, 2).value = entryList[i].transValue

root = tk.Tk()
frame = tk.Frame(root)
frame.width = 480
frame.height = 640
frame.pack()
categorise(ws)
w = tk.Label(root, text=sumcat(ws, 'Staff Costs'))
w2 = tk.Label(root, text=sumcat(ws, 'Equipment'))
w3 = tk.Label(root, text=sumcat(ws, 'Travel'))
w.pack()
w2.pack()
w3.pack()

from tkinter import filedialog
from tkinter import *
 
# root = Tk()
root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("excel files","*.xlsx"),("all files","*.*")))
print (root.filename)

messageButton = tk.Button(frame, text="Sum 101", command=sumcat(ws, 'Staff Costs'))
messageButton.pack(side=tk.LEFT)

# errmsg = 'Error!'
# mainloop()
root.mainloop()
# filename = input('File name? : ')
# output = input('Output file :') + '.xlsx'
# categorise(ws)
# deptlist = getList(ws, 1, 2)
# nominalList = getList(ws, 3, 4)
# writeList(ws, deptlist, 1, 1, 'Departments')
# writeList(ws, nominalList, 1, 1, 'Nominals')
# entryList = createEntries(ws)
# staffCosts = searchByNominal(entryList, 7003)
# writeEntryList(searchByNominal(entryList, 7003), 1, 'Entries')
# save(wb, output)
# print(sumcat(ws, 'Staff Costs'))
# print(sumdept(ws, 101))
# printList(deptlist)
# printList(nominalList)



