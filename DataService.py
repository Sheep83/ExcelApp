from openpyxl import load_workbook
# from Window import *
from Posting import *
from ListItem import *


class DataService():
    # def __init__(self):
        # self.wb = workbook

    def printSheetTitles(self, wb):
        for sheet in wb:
            print(sheet.title)

    def getRow(self, ws, rowno):
        # ws = wb[sheetName]
        newRow = ws[rowno]
        return newRow

    def importEntries(self, wb, sheetName):
        ws = wb[sheetName]        
        entryList = []
        for i in range (1, ws.max_row):
            newEntry = Posting(self.getRow(ws, i))
            entryList.append(newEntry)
        print(entryList.__len__())
        return entryList

    def categorise(self, ws):
        for i in range(1, ws.max_row):
            cell = ws.cell(i, 3)
            if (cell.value == 7003) or (cell.value == 7006) or (cell.value == 7007):
                ws.cell(i, 9, 'Staff Costs')
            if cell.value == 5032:
                ws.cell(i, 9, 'Equipment')
            if (cell.value == 7307) or (cell.value == 7311) or (cell.value == 7312):
                ws.cell(i, 9, 'Travel')

    def searchByNominal(self, entryList, nominalCode):
        filteredList = []
        for item in entryList:
            if item.nominalCode == nominalCode:
                filteredList.append(item)
        return filteredList

    def getList(self, ws, codeCol, nameCol):
        found = False
        list = []
        initCode = ws.cell(1, codeCol).value
        initName = ws.cell(1, nameCol).value
        list.append(ListItem(initCode, initName))
        for i in range(2, ws.max_row):
            itemCode = ws.cell(i, codeCol).value
            itemName = ws.cell(i, nameCol).value
            for item in list:
                if item.code == itemCode:
                    found = True
                else:
                    found = False
            if found == False:
                newItem = ListItem(itemCode, itemName)
                list.append(newItem)
        return list

    def writeList(self, wb, list, row, column, sheetname):
        newWs = wb.create_sheet(sheetname)
        i = 0
        for item in list:
            newWs.cell(row + i, column).value = item.code
            newWs.cell(row + i, column + 1).value = item.name
            i += 1

    def writeEntryList(self, wb, entryList, col1, sheetname):
        newWs = wb.create_sheet(sheetname)
        for i in range(1, entryList.__len__()):
            newWs.cell(i, 1).value = entryList[i].nominalCode
            newWs.cell(i, 2).value = entryList[i].transValue

    def openFile(self, filename):
        wb = load_workbook(filename)
        return wb

    def sumDept(self, dept, ws):
        deptTotal = 0
        for i in range(1, ws.max_row):
            value = ws.cell(i, 2).value
            if value == dept:
                deptTotal += ws.cell(i, 8).value
        return deptTotal





    
    